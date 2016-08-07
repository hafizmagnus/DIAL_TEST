# ---------------------
# This script was developed and authored by Hafiz Magnus for the exclusive use by RySense Ltd.
# Exclusive rights to the code structure and computational logic contained herein remains with the author.
# All data inputs/outputs are solely owned by RySense Ltd.
# ---------------------


import os, sys
import xlrd, xlsxwriter
import csv
import itertools
import openpyxl
from Tkinter import Tk
from tkFileDialog import askdirectory
import tkMessageBox
import glob
from xlsxwriter.workbook import Workbook
from openpyxl.chart import LineChart, Reference, Series


#---------------------
# function to convert MS Excel file to a CSV for processing
def csv_from_excel(xl_doc, xl_sheet, csv_file):
    wbc = xlrd.open_workbook(xl_doc)
    sh = wbc.sheet_by_name(xl_sheet)
    your_csv_file = open(csv_file, 'wb')
    wr = csv.writer(your_csv_file, quoting=csv.QUOTE_ALL)

    for rownum in xrange(sh.nrows):
        wr.writerow(sh.row_values(rownum))

    your_csv_file.close()


# function to convert seconds to minutes and seconds
def min2sec(seconds):
    if seconds or seconds.strip():
        if type(seconds) != int:
            seconds = int(seconds.strip())
        m,s = divmod(seconds, 60)
        h,m = divmod(m, 60)
        m = str(int(m)).zfill(2)
        s = str(int(s))
        if h > 0:
            h = int(h)
            h = str(h)
            h = h.zfill(2)
            return h + ":" + m + ":" + s
        else:
            return m + ":" + s
    else:
        pass

# function to populate list of vids, scale, and demographic questions
def vlist_pop(c_type):
    global v_lst
    try:
        c_type = c_type.upper()
        if c_type[0] == "M" and c_type not in v_lst:
            v_lst.append(c_type)
    except:
        pass

def slist_pop(c_type):
    global s_lst
    try:
        c_type = c_type.upper()
        if c_type[0] == "S" and c_type not in s_lst:
            s_lst.append(c_type)
    except:
        pass

def dlist_pop(c_type):
    global d_lst
    try:
        c_type = c_type.upper()
        if c_type[0] == "D" and c_type not in d_lst:
            d_lst.append(c_type)
    except:
        pass

"""function to find the row average"""
def row_average(f_row, video, values):
    global r_ave, val_lst
    if f_row.startswith(video):
        try:
            for item in itertools.chain(values[2:]):
                a = float(item)
                if a < 9999:
                    val_lst.append(a)
            r_tot = float(sum(val_lst))
            r_len = float(len(val_lst))
            r_ave = r_tot / r_len
            return r_ave
        except:
                pass

"""function to append items to list"""
def list_appender(*args):
    global f_write
    for arg in args:
        if isinstance(arg, list):
            for item in itertools.chain(arg):
                if item == None:
                    f_write.append("")
                else:
                    f_write.append(item)
        else:
            f_write.append(arg)
    return f_write

"""function to count the number of rows and return a list of respondens and add the average"""
def head_writer(filename, filename2):

    head_lst = ["Category", "Time"]
    respondents = []
    with open(filename, "r") as f:
        reader = csv.reader(f, lineterminator = '\n')
        first_row = next(reader)
        num_cols = len(first_row)
        f_col = num_cols - 2
        for x in range(f_col):
            i = x + 1
            i = "Respondent " + str(i)
            head_lst.append(i)
    head_lst.append("Dail Testing Score")
    head_lst.append("Mean Score")
    head_lst.append("Male_R_Average")
    head_lst.append("Male_C_Average")
    head_lst.append("Female_R_Average")
    head_lst.append("Female_C_Average")
    head_lst.append("Under_40_R_Average")
    head_lst.append("Under_40_C_Average")
    head_lst.append("Above_40_R_Average")
    head_lst.append("Above_40_C_Average")
    head_lst.append("Lower_Ed_R_Average")
    head_lst.append("Lower_Ed_C_Average")
    head_lst.append("Higher_Ed_R_Average")
    head_lst.append("Higher_Ed_C_Average")

    with open(filename2, "w") as r:
        writer = csv.writer(r, lineterminator = '\n')
        writer.writerow(head_lst)

"""function to populate the D1 dictionary"""
def position_dictionary(fist_csv_filename):
    global D1_dict
    with open(fist_csv_filename) as read:
        rdr = csv.reader(read, lineterminator = '\n')
        for row in rdr:
            if row[0] == "D1":
                gender = row[2:]
                position = []
                for i in range(len(gender)):
                    i = i + 2
                    position.append(i)
                D1_dict = dict(zip(position, gender))

    del(position)
    del(gender)

"""function to populate the D2 dictionary"""
def age_dictionary(fist_csv_filename):
    global D2_dict
    with open(fist_csv_filename) as read:
        rdr = csv.reader(read, lineterminator = '\n')
        for row in rdr:
            if row[0] == "D2":
                proto_list = row[2:]
                final_list = list()
                for item in proto_list:
                    if item == "1.0" or item == "2.0":
                        a = "1.0"
                    elif item == "3.0" or item == "4.0" or item == "5.0":
                        a = "2.0"
                    else:
                        a = "3.0"
                    final_list.append(a)
                position = []
                for i in range(len(final_list)):
                    i = i + 2
                    position.append(i)
                D2_dict = dict(zip(position, final_list))

    del(position)
    del(final_list)
    del(proto_list)

"""function to populate the D3 dictionary"""
def proto_ed_dic(fist_csv_filename):
    global protoD3_dict
    with open(fist_csv_filename) as read:
        rdr = csv.reader(read, lineterminator = '\n')
        for row in rdr:
            if row[0] == "D3":
                edu = row[2:]
                position = []
                for i in range(len(edu)):
                    i = i + 2
                    position.append(i)
                protoD3_dict = dict(zip(position, edu))

    del(position)
    del(edu)

def ed_dictionary(fist_csv_filename, age_dict, proto_list):
    global D3_dict
    with open(fist_csv_filename) as read:
        rdr = csv.reader(read, lineterminator = '\n')
        for row in rdr:
            final_list = list()
            if row[0] == "D3":                
                position = []
                for i in range(len(proto_list)):
                    i = i + 2
                    position.append(i)
                for item in position:
                    if age_dict[item] == "1.0" and proto_list[item] == "1.0":
                        a = "1.0"
                    elif age_dict[item] == "1.0" and proto_list[item] == "2.0":
                        a = "1.0"
                    elif age_dict[item] == "1.0" and proto_list[item] == "3.0":
                        a = "1.0"
                    elif age_dict[item] == "1.0" and proto_list[item] == "4.0":
                        a = "2.0"
                    elif age_dict[item] == "1.0" and proto_list[item] == "5.0":
                        a = "2.0"
                    elif age_dict[item] == "2.0" and proto_list[item] == "1.0":
                        a = "1.0"
                    elif age_dict[item] == "2.0" and proto_list[item] == "2.0":
                        a = "1.0"
                    elif age_dict[item] == "2.0" and proto_list[item] == "3.0":
                        a = "2.0"
                    elif age_dict[item] == "2.0" and proto_list[item] == "4.0":
                        a = "2.0"
                    elif age_dict[item] == "2.0" and proto_list[item] == "5.0":
                        a = "2.0"
                    else:
                        a = "3.0"
                    final_list.append(a)
                D3_dict = dict(zip(position, final_list))

    del(position)
    del(final_list)
    
"""function to get row average by gender"""
def gender_average(row, first, POS_DICT):
    global male_r_ave, female_r_ave
    male_total = 0
    female_total = 0
    male_count = 0
    female_count = 0
    if first.startswith("M"):
        p = 1
        for r in row[2:]:
            p += 1
            if POS_DICT[p] == '1.0':
                male_total += float(row[p])
                male_count += 1
            elif POS_DICT[p] == '2.0':
                female_total += float(row[p])
                female_count += 1
        male_r_ave = male_total / male_count
        female_r_ave = female_total / female_count

def noner(first):
    global male_r_ave, female_r_ave
    if not first.startswith("M"):
        male_r_ave = None
        female_r_ave = None

"""function to get row average by age"""
def age_average(row, first, POS_DICT):
    global u40_r_ave, a40_r_ave
    u40_total = 0
    a40_total = 0
    u40_count = 0
    a40_count = 0
    if first.startswith("M"):
        p = 1
        for r in row[2:]:
            p += 1
            if POS_DICT[p] == '1.0':
                u40_total += float(row[p])
                u40_count += 1
            elif POS_DICT[p] == '2.0':
                a40_total += float(row[p])
                a40_count += 1
        u40_r_ave = u40_total / u40_count
        a40_r_ave = a40_total / a40_count

def noner2(first):
    global u40_r_ave, a40_r_ave
    if not first.startswith("M"):
        u40_r_ave = None
        a40_r_ave = None

"""function to get row average by age"""
def edu_average(row, first, POS_DICT):
    global lower_ed_ave, higher_ed_ave
    l_total = 0
    h_total = 0
    l_count = 0
    h_count = 0
    if first.startswith("M"):
        p = 1
        for r in row[2:]:
            p += 1
            if POS_DICT[p] == '1.0':
                l_total += float(row[p])
                l_count += 1
            elif POS_DICT[p] == '2.0':
                h_total += float(row[p])
                h_count += 1
        lower_ed_ave = l_total / l_count
        higher_ed_ave = h_total / h_count

def noner3(first):
    global lower_ed_ave, higher_ed_ave
    if not first.startswith("M"):
        lower_ed_ave = None
        higher_ed_ave = None


"""function to count the number of MS Excel files processes and create a popup"""
def file_counter(folder_name):
    file_c = 0
    for croot, cdirs, cfiles in os.walk(folder_name):
        for cfile in cfiles:
            if cfile.endswith(".xls") or cfile.endswith(".xlsx"):
                file_c += 1

    tkMessageBox.showinfo(title="Success!", message = "A total of {} file(s) were processed.".format(file_c))

#---------------------
#getting user input for the root directory
Tk().withdraw()
base_dir = askdirectory(title = "Select folder with the Excel Workbooks for Dial Testing Analysis")


#mining for excel workbooks within the base directory
for root, dirs, files in os.walk(base_dir):
    for afile in files:
        if afile.endswith('.xlsx') or afile.endswith('.xls'):
            #initialising variables
            h_var = None
            s_lst = []
            v_lst = []
            d_lst = []
            f_write = []
            D1_dict = dict()
            D2_dict = dict()
            protoD3_dict = dict()
            D3_dict = dict()
            xl_path = os.path.join(base_dir, afile)
            csv_dir = base_dir + "\\" + "CSV"
            if not os.path.exists(csv_dir):
                os.makedirs(csv_dir)
            csv_path = os.path.join(csv_dir, "Temp.csv")
            temp2csv = os.path.join(csv_dir, "Temp2.csv")
            csv_from_excel(xl_path, "PA", csv_path)
            ana_dir = base_dir + "\\" + "Analysed"
            if not os.path.exists(ana_dir):
                os.makedirs(ana_dir)
            output_xl = os.path.join(ana_dir, ("Analysed_" + afile))


            head_writer(csv_path, temp2csv)
            position_dictionary(csv_path)
            age_dictionary(csv_path)
            proto_ed_dic(csv_path)
            ed_dictionary(csv_path, D2_dict, protoD3_dict)
            
            

            with open(csv_path) as ori:
                reader = csv.reader(ori, lineterminator = '\n')
                next(reader, None)
                for row in reader:
                    r = []
                    for item in row:
                        if item == "-1.0":
                            a = 9999
                            r.append(a)
                        else:
                            r.append(item)
                    f_write = []
                    dlist_pop(r[0])
                    vlist_pop(r[0])
                    slist_pop(r[0])
                    if r[0] != '':
                        h_var = r[0]
                    else:
                        pass
                    with open(temp2csv, "a") as ori2:
                        wrt = csv.writer(ori2, lineterminator = '\n')

                        #calculate the average row score per second
                        r_ave = None
                        val_lst = []
                        row_average(h_var, "M", r)

                        #calculate the average row score by gender per second
                        male_r_ave = 0
                        female_r_ave = 0
                        gender_average(r, h_var, D1_dict)
                        noner(h_var)

                        #calculate the average row score by age per second
                        u40_r_ave = 0
                        a40_r_ave = 0
                        age_average(r, h_var, D2_dict)
                        noner2(h_var)

                        #calculate the average row score by education level per second
                        lower_ed_ave = 0
                        higher_ed_ave = 0
                        edu_average(r, h_var, D3_dict)
                        noner3(h_var)

                        #create a list from all of the inputs and write the row in the csv
                        list_appender(h_var, min2sec(r[1]), r[2:], r_ave,"" , male_r_ave, "", female_r_ave, "", u40_r_ave, "", a40_r_ave, "", lower_ed_ave, "", higher_ed_ave, "")
                        wrt.writerow(f_write)

            #creating an individual file for each video
            for vid in v_lst:
                c_total = 0
                c_count = 0
                m_c_total = 0
                f_c_total = 0
                u40_c_total = 0
                a40_c_total = 0
                l_ed_c_total = 0
                h_ed_c_total = 0
                vid_out = os.path.join(csv_dir, (vid + ".csv"))

                #calculation of the overall average for each category
                with open(temp2csv, "r") as new:
                    rdr = csv.reader(new, lineterminator = '\n')
                    for row in rdr:
                        if row[0] == vid:
                            c_total += float(row[-14])
                            c_count += 1
                            m_c_total += float(row[-12])
                            f_c_total += float(row[-10])
                            u40_c_total += float(row[-8])
                            a40_c_total += float(row[-6])
                            l_ed_c_total += float(row[-4])
                            h_ed_c_total += float(row[-2])

                c_average = c_total / c_count
                m_c_average = m_c_total / c_count
                f_c_average = f_c_total / c_count
                u40_c_average = u40_c_total / c_count
                a40_c_average = a40_c_total / c_count
                l_ed_c_average = l_ed_c_total / c_count
                h_ed_c_average = h_ed_c_total / c_count


                with open(temp2csv, "r") as new2:
                    rdr = csv.reader(new2, lineterminator = '\n')
                    for row in rdr:
                        f_write = []
                        if row[0] == "Category":
                            with open(vid_out, "w") as new_vid:
                                wrt = csv.writer(new_vid, lineterminator = '\n')
                                list_appender(row)
                                wrt.writerow(f_write)
                        elif row[0] in d_lst:
                            with open(vid_out, "a") as new_vid:
                                wrt = csv.writer(new_vid, lineterminator = '\n')
                                list_appender(row)
                                wrt.writerow(f_write)
                        elif row[0] == vid:
                            with open(vid_out, "a") as new_vid:
                                wrt = csv.writer(new_vid, lineterminator = '\n')
                                list_appender(row[:-13], c_average, row[-12], m_c_average, row[-10], f_c_average, row[-8], u40_c_average, row[-6], a40_c_average, row[-4], l_ed_c_average, row[-2], h_ed_c_average)
                                wrt.writerow(f_write)


            os.remove(csv_path)
            os.remove(temp2csv)

#creating the final excel outputs without charts
            for broot, bdirs, bfiles in os.walk(csv_dir):
                for bfile in bfiles:
                    if bfile.endswith(".csv"):

                        head,tail = os.path.splitext(bfile) # <- getting the worksheetname from the .csv file
                        f_csv = os.path.join(broot, bfile) # <- CSV filename

                        if os.path.isfile(output_xl) == False:
                            wb = openpyxl.Workbook()
                            ws = wb.create_sheet()
                            ws.title = head

                            with open(f_csv, 'rb') as b_csv_file:
                                bf_reader = csv.reader(b_csv_file, lineterminator = "\n")
                                for row_index, row in enumerate(bf_reader):
                                    for column_index, cell in enumerate(row):
                                        column_letter = openpyxl.cell.get_column_letter((column_index + 1))
                                        s = cell
                                        try:
                                            s = float(s)
                                        except ValueError:
                                            pass
                                        ws.cell('%s%s'%(column_letter, (row_index + 1))).value = s



                            wb.save(output_xl)

                            try:
                                wb2 = openpyxl.load_workbook(output_xl)
                                ws2 = wb2.get_sheet_by_name("Sheet")
                                wb2.remove_sheet(ws2)
                                wb2.save(output_xl)
                            except:
                                pass


                        else:
                            wb = openpyxl.load_workbook(output_xl)
                            ws = wb.create_sheet()
                            ws.title = head

                            with open(f_csv, 'rb') as b_csv_file:
                                bf_reader = csv.reader(b_csv_file, lineterminator = "\n")
                                for row_index, row in enumerate(bf_reader):
                                    for column_index, cell in enumerate(row):
                                        column_letter = openpyxl.cell.get_column_letter((column_index + 1))
                                        s = cell
                                        try:
                                            s = float(s)
                                        except ValueError:
                                            pass
                                        ws.cell('%s%s'%(column_letter, (row_index + 1))).value = s
                            wb.save(output_xl)
                            try:
                                wb2 = openpyxl.load_workbook(output_xl)
                                ws2 = wb2.get_sheet_by_name("Sheet")
                                wb2.remove_sheet(ws2)
                                wb2.save(output_xl)
                            except:
                                pass

                        os.remove(f_csv)

os.removedirs(csv_dir)

from openpyxl.chart import LineChart, Reference, Series

#creating the necessary charts in each newly created MS Excel file
for croot, cdirs, cfiles in os.walk(ana_dir):
    for cfile in cfiles:
        if cfile.endswith(".xlsx"):
            cfile_name = os.path.join(ana_dir, cfile)
            cwb = openpyxl.load_workbook(cfile_name)
            ws_list = cwb.get_sheet_names()
            for dws in ws_list:
                cws = cwb.get_sheet_by_name(dws)
                col_max = cws.max_column
                row_max = int(cws.max_row)
                mm_cs = cwb.create_chartsheet()
                mm_cs.title = "Chart_" + dws + "_MM"
                
                DT_col = int(col_max) - 13                
                AC_col = int(col_max) - 12
                
                
                timings = Reference(cws, min_col = 2, min_row = 5, max_row = row_max)
                
                mm_r_ave = Reference(cws, min_col = DT_col, min_row = 5, max_row = row_max)
                mm_c_ave = Reference(cws, min_col = AC_col, min_row = 5, max_row = row_max)
                mm_chart = LineChart()
                mm_r_ave_ser = Series(mm_r_ave, title = "Dial Testing Score")
                mm_c_ave_ser = Series(mm_c_ave, title = "Mean Score")
                mm_chart.append(mm_r_ave_ser)
                mm_chart.append(mm_c_ave_ser)
                mm_chart.set_categories(timings)                
                mm_cs.add_chart(mm_chart)
                
                
                gender_cs = cwb.create_chartsheet()
                gender_cs.title = "Chart_" + dws + "_SEX"
                male_r_col = int(col_max) - 11
                male_ave = int(col_max) - 10
                female_r_col = int(col_max) - 9
                female_ave = int(col_max) - 8
                gender_male_r_ave = Reference(cws, min_col = male_r_col, min_row = 5, max_row = row_max)
                gender_female_r_ave = Reference(cws, min_col = female_r_col, min_row = 5, max_row = row_max)
                gender_chart = LineChart()
                gender_male_r_ave_ser = Series(gender_male_r_ave, title = "Male")
                gender_female_r_ave_ser = Series(gender_female_r_ave, title = "Female")
                gender_chart.append(gender_male_r_ave_ser)
                gender_chart.append(gender_female_r_ave_ser)
                gender_chart.append(mm_r_ave_ser)
                gender_chart.set_categories(timings)
                gender_cs.add_chart(gender_chart)
                
                age_cs = cwb.create_chartsheet()
                age_cs.title = "Chart_" + dws + "_AGE"
                u40_r_col = int(col_max) - 7
                a40_r_col = int(col_max) - 5
                age_u40_r_col = Reference(cws, min_col = u40_r_col, min_row = 5, max_row = row_max)
                age_a40_r_col = Reference(cws, min_col = a40_r_col, min_row = 5, max_row = row_max)
                age_chart = LineChart()
                age_u40_r_col_ser = Series(age_u40_r_col, title = "Younger")
                age_a40_r_col_ser = Series(age_a40_r_col, title = "Older")
                age_chart.append(age_u40_r_col_ser)
                age_chart.append(age_a40_r_col_ser)
                age_chart.append(mm_r_ave_ser)
                age_chart.set_categories(timings)
                age_cs.add_chart(age_chart)

                edu_cs = cwb.create_chartsheet()
                edu_cs.title = "Chart_" + dws + "_EDU"
                lower_r_ave = int(col_max) - 3
                higher_r_ave = int(col_max) - 1
                edu_lower_r_ave = Reference(cws, min_col = lower_r_ave, min_row = 5, max_row = row_max)
                edu_higher_r_ave = Reference(cws, min_col = higher_r_ave, min_row = 5, max_row = row_max)
                edu_chart = LineChart()
                edu_lower_r_ave_ser = Series(edu_lower_r_ave, title = "Lower Ed")
                edu_higher_r_ave_ser = Series(edu_higher_r_ave, title = "Higher Ed")
                edu_chart.append(edu_lower_r_ave_ser)
                edu_chart.append(edu_higher_r_ave_ser)
                edu_chart.append(mm_r_ave_ser)
                edu_chart.set_categories(timings)
                edu_cs.add_chart(edu_chart)
                
            cwb.save(cfile_name)
            
file_counter(ana_dir)
